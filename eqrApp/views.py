from django.shortcuts import redirect, render
import json
from django.contrib import messages
from django.contrib.auth.models import User
from django.http import HttpResponse
from eqrApp import models, forms
from django.db.models import Q
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl import Workbook
wb = Workbook()
from django.http import HttpResponseBadRequest
import pandas as pd
from django.core.files.storage import FileSystemStorage


def context_data():
    context = {
        'page_name' : '',
        'page_title' : 'Chat Room',
        'system_name' : 'Uni Design',
        'company_name_hindi' : 'सीप्ज़ ​​सेज़ अंधेरी (ई), मुंबई-96',
        'company_name_english' : 'SEEPZ SEZ Andheri (E), Mumbai-96',
        'topbar' : True,
        'footer' : True,
    }
    
    return context

# Create your views here.
def login_page(request):
    context = context_data()
    context['topbar'] = False
    context['footer'] = False
    context['page_name'] = 'login'
    context['page_title'] = 'Login'
    return render(request, 'login.html', context)

def login_user(request):
    logout(request)
    resp = {"status":'failed','msg':''}
    username = ''
    password = ''
    if request.POST:
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                resp['status']='success'
            else:
                resp['msg'] = "Incorrect username or password"
        else:
            resp['msg'] = "Incorrect username or password"
    return HttpResponse(json.dumps(resp),content_type='application/json')

@login_required
def home(request):
    context = context_data()
    context['page'] = 'home'
    context['page_title'] = 'Home'
    context['employees'] = models.Employee.objects.count()
    return render(request, 'home.html', context)

def logout_user(request):
    logout(request)
    return redirect('login-page')

@login_required
def employee_list(request):
    context =context_data()
    context['page'] = 'employee_list'
    context['page_title'] = 'Employee List'
    context['employees'] = models.Employee.objects.all()
    return render(request, 'employee_list.html', context)


@login_required 
def manage_employee(request, pk=None):
    context =context_data()
    if pk is None:
        context['page'] = 'add_employee'
        context['page_title'] = 'Add New Employee'
        context['employee'] = {}
    else:
        context['page'] = 'edit_employee'
        context['page_title'] = 'Update Employee'
        context['employee'] = models.Employee.objects.get(id=pk)
    return render(request, 'manage_employee.html', context)

@login_required
def save_employee(request):
    resp = { 'status' : 'failed', 'msg' : '' }
    if not request.method == 'POST':
        resp['msg'] = "No data has been sent into the request."

    else:
        if request.POST['id'] == '':
            form = forms.SaveEmployee(request.POST, request.FILES)
        else:
            employee = models.Employee.objects.get(id = request.POST['id'])
            form = forms.SaveEmployee(request.POST, request.FILES, instance = employee)
        if form.is_valid():
            form.save()
            if request.POST['id'] == '':
                messages.success(request, f"{request.POST['employee_code']} has been added successfully.")
            else:
                messages.success(request, f"{request.POST['employee_code']} has been updated successfully.")
            resp['status'] = 'success'
        else:
            for field in form:
                for error in field.errors:
                    if not resp['msg'] == '':
                        resp['msg'] += str("<br />")
                    resp['msg'] += str(f"[{field.label}] {error}")
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def save_employee_excel(request):
    # if request.method == "POST":
    #     excel_file = request.FILES["file"]
    #     wb = openpyxl.load_workbook(excel_file)
    #     worksheet = wb["Sheet1"]
    #     excel_data = []
        
    #     for row in worksheet.iter_rows(values_only=True):
    #         employee_code, first_name, middle_name, last_name, dob, gender, contact, email, address, department, position, avatar = row[:12]

    #         form = forms.SaveEmployee(data={
    #             'employee_code': employee_code,
    #             'first_name': first_name,
    #             'middle_name': middle_name,
    #             'last_name': last_name,
    #             'dob': dob,
    #             'gender': gender,
    #             'contact': contact,
    #             'email': email,
    #             'address': address,
    #             'department': department,
    #             'position': position,
    #             'avatar': avatar
    #         })
            
    #         if form.is_valid():
    #             form.save()
    #             messages.success(request, "Employee has been added successfully.")
    #         else:
    #             print(form.errors)
    #             messages.error(request, "Data Not Inserted. Please try again Testing...")

    if request.method == 'POST' and request.FILES.get('file'):
        myfile = request.FILES['file']
        if myfile.name.endswith('.xlsx'):
            try:
                empexceldata = pd.read_excel(myfile)
                
                # List to keep track of duplicate employees
                duplicate_employees = []

                for row in empexceldata.itertuples():
                    # Check if similar data already exists
                    existing_record = models.Employee.objects.filter(
                        employee_code=row.employee_code,
                        first_name=row.first_name,
                        middle_name=row.middle_name,
                        last_name=row.last_name,
                        dob=row.dob,
                        gender=row.gender,
                        contact=row.contact,
                        email=row.email,
                        address=row.address,
                        department=row.department,
                        position=row.position,
                        avatar=row.avatar
                    ).exists()

                    # If similar data exists, record it
                    if existing_record:
                        duplicate_employees.append(row)

                    # If similar data does not exist, insert new data
                    else:
                        obj = models.Employee(
                            employee_code=row.employee_code,
                            first_name=row.first_name,
                            middle_name=row.middle_name,
                            last_name=row.last_name,
                            dob=row.dob,
                            gender=row.gender,
                            contact=row.contact,
                            email=row.email,
                            address=row.address,
                            department=row.department,
                            position=row.position,
                            avatar=row.avatar
                        )
                        obj.save()
                
                fs = FileSystemStorage()
                filename = fs.save(myfile.name, myfile)
                uploaded_file_url = fs.url(filename)
                if not duplicate_employees:
                        # Add success message if no duplicate employees found
                            messages.success(request, "Employee data has been processed successfully.")
                        # Render the page
                            return render(request, 'add_employee_excel.html', {'message': messages})
            
                # If there are duplicate employees, inform the user
                else:
                    messages.warning(request, f"The following employees already exist: {', '.join([f'{row.first_name} {row.last_name}' for row in duplicate_employees])}")

                return render(request, 'add_employee_excel.html', {'message': messages})
            
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
                return render(request, 'add_employee_excel.html', {'message': messages})
            else:
                messages.error(request, "Uploaded file is not in Excel format.")
            return render(request, 'add_employee_excel.html', {'message': messages})
        else:
            return HttpResponseBadRequest("No file uploaded or invalid request method.")

@login_required
def view_card(request, pk =None):
    if pk is None:
        return HttpResponse("Employee ID is Invalid")
    else:
        context = context_data()
        context['employee'] = models.Employee.objects.get(id=pk)
        return render(request, 'view_id.html', context)

@login_required
def view_scanner(request):
    context = context_data()
    return render(request, 'scanner.html', context)

@login_required
def add_employee_excel(request):
        context =context_data()
        context['page_title'] = 'Upload Employee Data'
        return render(request, 'add_employee_excel.html', context)
    
def   id_testing (request):
        # context =context_data()
        # context['page_title'] = 'Employee ID CARD'
        # context = context_data()
        # context['employee_code'] = models.Employee.objects.get(employee_code=code)
        return render(request, 'id_testing.html')
    
@login_required
def view_details(request, code = None):
    if code is None:
        return HttpResponse("Employee code is Invalid")
    else:
        context = context_data()
        context['employee'] = models.Employee.objects.get(employee_code=code)
        return render(request, 'view_details.html', context)

@login_required
def delete_employee(request, pk=None):
    resp = { 'status' : 'failed', 'msg' : '' }
    if pk is None:
        resp['msg'] = "No data has been sent into the request."
    else:
        try:
            models.Employee.objects.get(id=pk).delete()
            resp['status'] = 'success'
            messages.success(request, 'Employee has been deleted successfully.')
        except:
            resp['msg'] = "Employee has failed to delete."

    return HttpResponse(json.dumps(resp), content_type="application/json")
